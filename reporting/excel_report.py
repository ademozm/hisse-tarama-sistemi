"""
Excel rapor oluşturma v2.

Sayfalar:
- Özet          : En iyi 20 al / en iyi 20 sat adayı (filtrelenmemiş, tüm sinyaller)
- Filtrelenmiş  : config.FILTERS (veya CLI override) kriterlerinden geçen sonuçlar
- ABD, BIST, Kripto : Piyasa bazlı TAM tarama sonuçları (filtrelenmemiş)
- Temel Analiz  : Ham P/E, ROE, borç, büyüme verileri (sadece hisseler)
- Risk Metrikleri : Volatilite, max drawdown, 52 haftalık aralık, beta
- Filtre Özeti  : Kaç sembolün hangi filtrede elendiğinin dökümü
- Hata Raporu   : İndirilemeyen veya doğrulamadan geçemeyen semboller

Her sayfada: donmuş başlık satırı, otomatik filtre, skor kolonunda renk skalası.
"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

DISPLAY_COLUMNS = [
    "symbol", "name", "market", "signal", "regime", "composite_score",
    "close", "momentum_return_pct", "relative_strength_pct", "risk_reward",
    "atr_pct", "adx", "rsi", "fundamental_score", "relative_volume",
    "volume_confirmed", "mtf_confirmed", "stooq_close", "fark_yuzde", "supheli",
    "onerilen_adet", "pozisyon_buyuklugu", "portfoy_yuzdesi",
]

COLUMN_LABELS = {
    "symbol": "Sembol", "name": "Şirket/Varlık", "market": "Piyasa",
    "signal": "Sinyal", "regime": "Rejim", "composite_score": "Skor",
    "close": "Kapanış", "momentum_return_pct": "Momentum %",
    "relative_strength_pct": "Göreceli Güç %", "risk_reward": "Risk/Ödül",
    "atr_pct": "ATR %", "adx": "ADX", "rsi": "RSI",
    "fundamental_score": "Temel Skor", "relative_volume": "Bağıl Hacim",
    "volume_confirmed": "Hacim Teyidi", "mtf_confirmed": "Haftalık Teyit",
    "stooq_close": "Stooq Kapanış (2. Kaynak)", "fark_yuzde": "Kaynak Farkı %",
    "supheli": "Veri Şüpheli mi",
    "onerilen_adet": "Önerilen Adet", "pozisyon_buyuklugu": "Pozisyon Büyüklüğü ($)",
    "portfoy_yuzdesi": "Portföy Yüzdesi %",
}

FUNDAMENTAL_COLUMNS = [
    "symbol", "name", "market", "trailingPE", "forwardPE", "priceToBook",
    "pegRatio", "returnOnEquity", "debtToEquity", "revenueGrowth",
    "earningsGrowth", "dividendYield", "marketCap", "profitMargins",
    "fundamental_score",
]
FUNDAMENTAL_LABELS = {
    "symbol": "Sembol", "name": "Şirket", "market": "Piyasa",
    "trailingPE": "F/K (Trailing)", "forwardPE": "F/K (Forward)",
    "priceToBook": "PD/DD", "pegRatio": "PEG", "returnOnEquity": "ROE",
    "debtToEquity": "Borç/Özkaynak", "revenueGrowth": "Gelir Büyümesi",
    "earningsGrowth": "Kâr Büyümesi", "dividendYield": "Temettü Verimi",
    "marketCap": "Piyasa Değeri", "profitMargins": "Kâr Marjı",
    "fundamental_score": "Temel Skor",
}

RISK_COLUMNS = [
    "symbol", "name", "market", "volatility_annualized_pct", "max_drawdown_pct",
    "week52_high", "week52_low", "pct_from_52w_high", "pct_from_52w_low",
    "week52_range_position", "beta",
]
RISK_LABELS = {
    "symbol": "Sembol", "name": "Şirket/Varlık", "market": "Piyasa",
    "volatility_annualized_pct": "Yıllık Volatilite %",
    "max_drawdown_pct": "Maks. Düşüş %", "week52_high": "52H Zirve",
    "week52_low": "52H Dip", "pct_from_52w_high": "Zirveden Uzaklık %",
    "pct_from_52w_low": "Dipten Uzaklık %",
    "week52_range_position": "52H Aralık Konumu (0=dip,1=zirve)", "beta": "Beta",
}

ADVANCED_COLUMNS = [
    "symbol", "name", "market", "signal", "close",
    "fib_trend_direction", "nearest_fib_level", "nearest_fib_distance_pct",
    "support_levels", "resistance_levels", "poc_price",
    "value_area_low", "value_area_high",
    "candlestick_pattern", "candlestick_direction",
]
ADVANCED_LABELS = {
    "symbol": "Sembol", "name": "Şirket/Varlık", "market": "Piyasa",
    "signal": "Sinyal", "close": "Kapanış",
    "fib_trend_direction": "Fib. Trend Yönü", "nearest_fib_level": "En Yakın Fib. Seviyesi",
    "nearest_fib_distance_pct": "Fib. Seviyesine Uzaklık %",
    "support_levels": "Destek Seviyeleri", "resistance_levels": "Direnç Seviyeleri",
    "poc_price": "POC (Hacim Odağı)", "value_area_low": "Değer Alanı Alt",
    "value_area_high": "Değer Alanı Üst",
    "candlestick_pattern": "Mum Formasyonu", "candlestick_direction": "Formasyon Yönü",
}


def _style_sheet(ws, ncols, score_col_name=None, headers=None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if score_col_name and headers and score_col_name in headers:
        col_idx = headers.index(score_col_name) + 1
        col_letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="min", start_color="F08080",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="90EE90",
        )
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)


def _safe_str_len(value) -> int:
    if pd.isna(value):
        return 0
    return len(str(value))


def _autosize(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        content_max = df[col_name].map(_safe_str_len).max()
        max_len = max(content_max if pd.notna(content_max) else 0, len(str(col_name))) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)


def _write_generic_sheet(writer, df: pd.DataFrame, sheet_name: str, columns, labels, score_col=None):
    if df.empty:
        pd.DataFrame({"Not": ["Bu sayfada gösterilecek veri bulunamadı"]}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    display_df = df[[c for c in columns if c in df.columns]].copy()
    if "signal" in display_df.columns:
        display_df["signal"] = display_df["signal"].map({1: "AL", -1: "SAT"})
    for bool_col in ("volume_confirmed", "mtf_confirmed", "supheli"):
        if bool_col in display_df.columns:
            # NOT: "N/A" KULLANMA — pandas Excel'den geri okurken "N/A" metnini
            # otomatik olarak eksik veri (NaN) sayıyor, bu da rapor tekrar
            # okunduğunda (örn. Streamlit panelinde) veri sessizce kayboluyor.
            display_df[bool_col] = display_df[bool_col].map({True: "Evet", False: "Hayır"}).fillna("Bilinmiyor")
    for list_col in ("support_levels", "resistance_levels"):
        if list_col in display_df.columns:
            display_df[list_col] = display_df[list_col].apply(
                lambda v: ", ".join(str(x) for x in v) if isinstance(v, list) and v else "-"
            )
    display_df = display_df.rename(columns=labels)
    display_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    score_label = labels.get(score_col) if score_col else None
    _style_sheet(ws, len(display_df.columns), score_col_name=score_label, headers=list(display_df.columns))
    _autosize(ws, display_df)


def build_report(
    scored_df: pd.DataFrame,
    validation_report: dict,
    output_path: str,
    filtered_df: pd.DataFrame | None = None,
    filter_stats: dict | None = None,
    performance_stats_df: pd.DataFrame | None = None,
    macro_news: list | None = None,
    calendar_df: pd.DataFrame | None = None,
    grid_plan_df: pd.DataFrame | None = None,
    dca_plan_df: pd.DataFrame | None = None,
    full_status_df: pd.DataFrame | None = None,
    grid_dca_performance: dict | None = None,
):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # --- Özet: en iyi 20 al / en iyi 20 sat (filtrelenmemiş tüm evrenden) ---
        if not scored_df.empty:
            top_buy = scored_df[scored_df["signal"] == 1].head(20)
            top_sell = scored_df[scored_df["signal"] == -1].sort_values("composite_score").head(20)
            summary = pd.concat([top_buy, top_sell])
        else:
            summary = scored_df
        _write_generic_sheet(writer, summary, "Özet", DISPLAY_COLUMNS, COLUMN_LABELS, score_col="composite_score")

        # --- Tüm Sembol Durumu: sinyal üretmeyenler DAHİL, taranan her sembolün
        # şeffaf durumu. "Piyasa boş/az görünüyor" kafa karışıklığını önlemek için. ---
        if full_status_df is not None and not full_status_df.empty:
            status_cols = ["symbol", "name", "market", "regime", "close", "adx", "rsi", "sinyal_var_mi", "neden"]
            status_labels = {
                "symbol": "Sembol", "name": "Şirket/Varlık", "market": "Piyasa",
                "regime": "Rejim", "close": "Kapanış", "adx": "ADX", "rsi": "RSI",
                "sinyal_var_mi": "Sinyal Var mı", "neden": "Neden",
            }
            status_display = full_status_df[[c for c in status_cols if c in full_status_df.columns]].copy()
            status_display["sinyal_var_mi"] = status_display["sinyal_var_mi"].map({True: "Evet", False: "Hayır"})
            status_display = status_display.rename(columns=status_labels)
            status_display.to_excel(writer, sheet_name="Tüm Sembol Durumu", index=False)
            ws = writer.sheets["Tüm Sembol Durumu"]
            _style_sheet(ws, len(status_display.columns))
            _autosize(ws, status_display)

            # ÖNEMLİ: Bu özet tablo BİLEREK AYRI bir sayfaya yazılıyor, "Tüm Sembol
            # Durumu" sayfasının altına DEĞİL. Aynı sayfaya alt alta iki tablo
            # yazmak Excel'de görsel olarak sorunsuz görünür ama pd.read_excel ile
            # programatik okunduğunda (örn. Streamlit panelinde) iki tablo birbirine
            # karışıp veriyi bozar — bu gerçek bir hataydı, testle yakalandı.
            market_summary = full_status_df.groupby("market").agg(
                taranan=("symbol", "count"), sinyal_ureten=("sinyal_var_mi", "sum")
            ).reset_index().rename(columns={"market": "Piyasa", "taranan": "Taranan Sembol",
                                             "sinyal_ureten": "Sinyal Üreten"})
            market_summary["Sinyal Üretme Oranı %"] = (
                market_summary["Sinyal Üreten"] / market_summary["Taranan Sembol"] * 100
            ).round(1)
            market_summary.to_excel(writer, sheet_name="Piyasa Özeti", index=False)
            ws2 = writer.sheets["Piyasa Özeti"]
            _style_sheet(ws2, len(market_summary.columns))
            _autosize(ws2, market_summary)
        else:
            pd.DataFrame({"Not": ["Bu taramada hiçbir sembol için veri işlenemedi."]}
                         ).to_excel(writer, sheet_name="Tüm Sembol Durumu", index=False)

        # --- Filtrelenmiş sonuçlar ---
        if filtered_df is not None:
            _write_generic_sheet(writer, filtered_df, "Filtrelenmiş", DISPLAY_COLUMNS, COLUMN_LABELS,
                                  score_col="composite_score")

        # --- Piyasa bazlı tam sonuçlar ---
        for market, label in [("us", "ABD"), ("bist", "BIST"), ("crypto", "Kripto"),
                               ("emtia", "Emtialar"), ("forex", "Döviz")]:
            market_df = scored_df[scored_df["market"] == market] if not scored_df.empty else scored_df
            _write_generic_sheet(writer, market_df, label, DISPLAY_COLUMNS, COLUMN_LABELS,
                                  score_col="composite_score")

        # --- Temel analiz detayı ---
        fundamental_df = scored_df[~scored_df["market"].isin(["crypto", "emtia", "forex"])] if not scored_df.empty else scored_df
        _write_generic_sheet(writer, fundamental_df, "Temel Analiz", FUNDAMENTAL_COLUMNS, FUNDAMENTAL_LABELS)

        # --- Gelişmiş göstergeler (Fibonacci, destek/direnç, hacim profili) ---
        _write_generic_sheet(writer, scored_df, "Gelişmiş Göstergeler", ADVANCED_COLUMNS, ADVANCED_LABELS)

        # --- Risk metrikleri ---
        _write_generic_sheet(writer, scored_df, "Risk Metrikleri", RISK_COLUMNS, RISK_LABELS)

        # --- Performans geçmişi (sinyal günlüğü) ---
        if performance_stats_df is not None and not performance_stats_df.empty:
            performance_stats_df.to_excel(writer, sheet_name="Performans Geçmişi", index=False)
            ws = writer.sheets["Performans Geçmişi"]
            _style_sheet(ws, len(performance_stats_df.columns))
            _autosize(ws, performance_stats_df)
        else:
            note_df = pd.DataFrame({"Not": [
                "Henüz kapanmış (kazandı/kaybetti sonuçlanmış) sinyal yok. "
                "Sistem birkaç tarama sonra bu sayfayı otomatik dolduracak."
            ]})
            note_df.to_excel(writer, sheet_name="Performans Geçmişi", index=False)

        # --- Haberler (sembol bazlı + ABD piyasasını genel etkileyen makro haberler) ---
        news_rows = []
        if not scored_df.empty and "latest_headline" in scored_df.columns:
            news_cols = ["symbol", "name", "market", "news_sentiment", "news_count", "latest_headline"]
            symbol_news_df = scored_df[[c for c in news_cols if c in scored_df.columns]].copy()
            symbol_news_df = symbol_news_df.dropna(subset=["latest_headline"])
            if not symbol_news_df.empty:
                symbol_news_df["kaynak"] = "Sembol haberi"
                symbol_news_df = symbol_news_df.rename(columns={
                    "symbol": "Sembol", "name": "Şirket/Varlık", "market": "Piyasa",
                    "news_sentiment": "Haber Tonu", "news_count": "Haber Sayısı",
                    "latest_headline": "Başlık", "kaynak": "Kaynak",
                })
                news_rows.append(symbol_news_df)

        if macro_news:
            macro_df = pd.DataFrame(macro_news)
            if not macro_df.empty:
                macro_df = macro_df.rename(columns={
                    "title": "Başlık", "publisher": "Yayıncı", "sentiment": "Haber Tonu", "kaynak": "Kaynak",
                })
                macro_df["Sembol"] = ""
                macro_df["Şirket/Varlık"] = ""
                macro_df["Piyasa"] = "makro"
                news_rows.append(macro_df)

        if news_rows:
            combined_news = pd.concat(news_rows, ignore_index=True)
            cols_order = [c for c in ["Kaynak", "Sembol", "Şirket/Varlık", "Piyasa", "Başlık", "Haber Tonu", "Yayıncı", "Haber Sayısı"] if c in combined_news.columns]
            combined_news = combined_news[cols_order]
            combined_news.to_excel(writer, sheet_name="Haberler", index=False)
            ws = writer.sheets["Haberler"]
            _style_sheet(ws, len(combined_news.columns), score_col_name="Haber Tonu", headers=list(combined_news.columns))
            _autosize(ws, combined_news)
        else:
            note_df = pd.DataFrame({"Not": [
                "Bu taramada haber verisi bulunamadı (--skip-news kullanılmış olabilir, "
                "ya da yfinance o an için haber döndürmedi)."
            ]})
            note_df.to_excel(writer, sheet_name="Haberler", index=False)


        # --- Ekonomik takvim (yaklaşan Fed toplantıları, önemli veri açıklamaları) ---
        if calendar_df is not None and not calendar_df.empty:
            calendar_display = calendar_df.rename(columns={
                "tarih": "Tarih", "olay": "Olay", "kalan_gun": "Kalan Gün", "onem": "Önem",
            })
            calendar_display.to_excel(writer, sheet_name="Ekonomik Takvim", index=False)
            ws = writer.sheets["Ekonomik Takvim"]
            _style_sheet(ws, len(calendar_display.columns))
            _autosize(ws, calendar_display)
        else:
            note_df = pd.DataFrame({"Not": [
                "Önümüzdeki 14 gün içinde bilinen önemli bir ekonomik olay (FOMC, NFP, CPI) yok."
            ]})
            note_df.to_excel(writer, sheet_name="Ekonomik Takvim", index=False)

        # --- Grid strateji planı (sadece yatay/range rejimindeki semboller) ---
        if grid_plan_df is not None and not grid_plan_df.empty:
            grid_display = grid_plan_df.rename(columns={
                "symbol": "Sembol", "seviye": "Seviye", "al_fiyati": "Al Fiyatı",
                "sat_fiyati": "Sat Fiyatı", "adet": "Adet", "beklenen_kar_pct": "Beklenen Kâr %",
            })
            grid_display.to_excel(writer, sheet_name="Grid Planı", index=False)
            ws = writer.sheets["Grid Planı"]
            _style_sheet(ws, len(grid_display.columns))
            _autosize(ws, grid_display)
        else:
            pd.DataFrame({"Not": ["Bu taramada yatay (range) rejiminde, grid stratejisine uygun sembol bulunamadı."]}
                         ).to_excel(writer, sheet_name="Grid Planı", index=False)

        # --- DCA (kademeli alım) planı ---
        if dca_plan_df is not None and not dca_plan_df.empty:
            dca_display = dca_plan_df.rename(columns={
                "symbol": "Sembol", "dilim": "Dilim", "tetik_fiyati": "Tetik Fiyatı",
                "fiyat_dususu_pct": "Fiyat Düşüşü %", "tutar": "Tutar ($)",
                "adet": "Adet", "kumulatif_tutar": "Kümülatif Tutar ($)",
            })
            dca_display.to_excel(writer, sheet_name="DCA Planı", index=False)
            ws = writer.sheets["DCA Planı"]
            _style_sheet(ws, len(dca_display.columns))
            _autosize(ws, dca_display)
        else:
            pd.DataFrame({"Not": ["Bu taramada AL sinyali üreten sembol bulunamadı, DCA planı oluşturulmadı."]}
                         ).to_excel(writer, sheet_name="DCA Planı", index=False)

        # --- Grid & DCA Performansı (SQLite emir günlüğünden gerçek kazanma oranları) ---
        grid_perf = (grid_dca_performance or {}).get("grid", {})
        dca_perf = (grid_dca_performance or {}).get("dca", {})

        grid_perf_rows = [
            {"Metrik": "Kapanan işlem (al+sat tamamlandı)", "Değer": grid_perf.get("kapanan_islem", 0)},
            {"Metrik": "Kazanma oranı %", "Değer": grid_perf.get("kazanma_orani_pct") or "Henüz yeterli veri yok"},
            {"Metrik": "Ortalama kazanç %", "Değer": grid_perf.get("ortalama_kazanc_pct") or "Henüz yeterli veri yok"},
            {"Metrik": "Bekleyen emir (tetiklenmeyi bekliyor)", "Değer": grid_perf.get("bekleyen", 0)},
            {"Metrik": "Süresi dolan emir (60+ gün tetiklenmedi)", "Değer": grid_perf.get("suresi_dolan", 0)},
        ]
        dca_perf_rows = [
            {"Metrik": "Gerçekleşen dilim", "Değer": dca_perf.get("gerceklesen_dilim", 0)},
            {"Metrik": "Bekleyen dilim", "Değer": dca_perf.get("bekleyen_dilim", 0)},
            {"Metrik": "Ortalama getiri % (güncel fiyata göre)", "Değer": dca_perf.get("ortalama_getiri_pct") or "Henüz yeterli veri yok"},
            {"Metrik": "Pozitif pozisyon oranı %", "Değer": dca_perf.get("pozitif_pozisyon_orani_pct") or "Henüz yeterli veri yok"},
        ]

        ws = writer.book.create_sheet("Grid ve DCA Performansı")
        ws["A1"] = "Grid Stratejisi Performansı (gerçekleşen emirlerden hesaplanır)"
        ws["A1"].font = Font(bold=True)
        grid_perf_df = pd.DataFrame(grid_perf_rows)
        grid_perf_df.to_excel(writer, sheet_name="Grid ve DCA Performansı", index=False, startrow=1)

        dca_start_row = len(grid_perf_rows) + 4
        ws[f"A{dca_start_row}"] = "DCA Stratejisi Performansı (gerçekleşen dilimlerden hesaplanır)"
        ws[f"A{dca_start_row}"].font = Font(bold=True)
        dca_perf_df = pd.DataFrame(dca_perf_rows)
        dca_perf_df.to_excel(writer, sheet_name="Grid ve DCA Performansı", index=False, startrow=dca_start_row)
        _autosize(ws, pd.concat([grid_perf_df, dca_perf_df], ignore_index=True))

        # --- Filtre özeti ---
        if filter_stats:
            stats_df = pd.DataFrame(list(filter_stats.items()), columns=["Aşama", "Kalan Sembol Sayısı"])
            stats_df.to_excel(writer, sheet_name="Filtre Özeti", index=False)
            ws = writer.sheets["Filtre Özeti"]
            _style_sheet(ws, 2)
            _autosize(ws, stats_df)

        # --- Hata raporu ---
        error_rows = [
            {"Sembol": sym, "Neden": "; ".join(vr.reasons)}
            for sym, vr in validation_report.items() if not vr.is_valid
        ]
        error_df = pd.DataFrame(error_rows) if error_rows else pd.DataFrame({"Not": ["Hata yok"]})
        error_df.to_excel(writer, sheet_name="Hata Raporu", index=False)

    return output_path
